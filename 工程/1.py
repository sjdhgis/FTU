import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

def copy_and_rename_docx_files(source_folder, target_folder, excel_file, sheet_name, column_name):
    """
    根据Excel文件名的数量复制和重命名源文件夹中的.docx文件。

    参数:
        source_folder (str): 源文件夹路径，包含需要复制的.docx文件。
        target_folder (str): 目标文件夹路径，复制后的文件将存储在此。
        excel_file (str): Excel文件路径，包含用于重命名的文件名。
        sheet_name (str): Excel工作表名称。
        column_name (str): 包含文件名的列名（例如'A'或'B'）。
    """
    # 检查源文件夹是否存在
    if not os.path.exists(source_folder) or not os.path.isdir(source_folder):
        print(f"错误：源文件夹 '{source_folder}' 不存在或不是一个文件夹。")
        return

    # 确保目标文件夹存在
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)

    # 加载Excel文件
    try:
        workbook = load_workbook(filename=excel_file)
        sheet = workbook[sheet_name]
    except Exception as e:
        print(f"错误：无法加载Excel文件或工作表。{e}")
        return

    # 获取指定列的索引（从0开始）
    try:
        column_index = column_index_from_string(column_name) - 1
    except ValueError:
        print(f"错误：列名 '{column_name}' 无效。")
        return

    # 获取指定列的所有值（跳过标题行）
    file_names = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行
        if len(row) > column_index:  # 确保列索引有效
            file_name = row[column_index]
            if file_name and isinstance(file_name, str):  # 确保文件名有效且为字符串
                file_names.append(file_name.strip())  # 去除可能的空白字符

    # 获取源文件夹中的所有.docx文件
    try:
        docx_files = [f for f in os.listdir(source_folder) if f.endswith('.docx')]
    except Exception as e:
        print(f"错误：无法列出源文件夹中的文件。{e}")
        return

    # 检查文件数量是否匹配
    if len(docx_files) > len(file_names):
        print("警告：源文件夹中的.docx文件数量多于Excel表格中的文件名数量。")
        print(f"源文件夹中有 {len(docx_files)} 个.docx文件，而Excel中有 {len(file_names)} 个文件名。")
        return

    # 如果源文件夹中的文件数量少于Excel文件名的数量，循环复制文件
    if len(docx_files) < len(file_names):
        print(f"源文件夹中的.docx文件数量少于Excel文件名数量，将循环复制文件以匹配数量。")
        docx_files = (docx_files * (len(file_names) // len(docx_files) + 1))[:len(file_names)]

    # 复制并重命名文件
    for docx_file, new_name in zip(docx_files, file_names):
        source_path = os.path.join(source_folder, docx_file)
        target_path = os.path.join(target_folder, f"{new_name}.docx")
        shutil.copy(source_path, target_path)
        print(f"文件 {docx_file} 已复制并重命名为 {new_name}.docx")

    print("所有文件已成功复制并重命名。")

# 示例用法
source_folder =r"C:\Users\33\Desktop\2024年网改项目结算\网改项目\项目\项目工程"   # 源文件夹路径
target_folder = r"C:\Users\33\Desktop\2024年网改项目结算\网改项目\项目\工程\1"  # 目标文件夹路径
excel_file = r"C:\Users\33\Desktop\2024年网改项目结算\网改项目\项目\模板.xlsx"  # Excel文件路径
sheet_name = "Sheet3"  # Excel工作表名称
column_name = "A"  # 包含文件名的列名

copy_and_rename_docx_files(source_folder, target_folder, excel_file, sheet_name, column_name)

# 示例用法
#source_folder = r"C:\Users\33\Desktop\2024年网改项目结算\网改项目\项目\文件"  # 源文件夹路径
#target_folder = r"C:\Users\33\Desktop\2024年网改项目结算\网改项目\项目\文件"  # 目标文件夹路径
#excel_file = r"C:\Users\33\Desktop\2024年网改项目结算\网改项目\项目\模板.xlsx"  # Excel文件路径
#sheet_name = "Sheet1"  # Excel工作表名称
#column_name = "G"  # 包含文件名的列名

#copy_and_rename_docx_files(source_folder, target_folder, excel_file, sheet_name, column_name)


